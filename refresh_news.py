# -*- coding: utf-8 -*-
"""refresh_news_auto.py

JEDEN skript, který pokaždé vytvoří NOVÝ Excel a naplní ho:
- tickery: z MT5 (viditelné symboly v MarketWatch)
- news: RSS podle šablon (zdroje jsou definované v kódu a zároveň se zapíšou do sheetu Sources)
- technika: počítá se z MT5 (D1) -> TechScore(0–50)
- Yahoo: yfinance -> YahooScore(-20..20) + detail sloupce
- MarketCap+Rank: volitelné (pokud existuje soubor market_watch_symbols_enriched_yahoo.xlsx)

Pozn:
- Použitá logika TotalScore + Signal je převzatá z tvé "refresh_news - kopie.py"
  (tj. STRONG BUY/BUY/HOLD/SELL/STRONG SELL).
- V Excelu NEZAPISUJEME timezone-aware datetime (Excel/openpyxl to neumí).

Spuštění:
  python refresh_news_auto.py

Volitelné argumenty:
  python refresh_news_auto.py --outdir "C:\\Users\\...\\Documents\\JOHNY" \
      --marketcap "market_watch_symbols_enriched_yahoo.xlsx"

"""

from __future__ import annotations

import argparse
import datetime as dt
import logging
import math
import os
import re
import sys
import time
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


# -----------------------------
# SOURCES (fundamentální / news)
# -----------------------------

SOURCES = [
    {"source": "Česká národní banka (OAM + krátké pozice)", "type": "regulator filings", "paywall": "free", "info_level": 5, "template": "", "notes": "S1: veřejné OAM/krátké pozice, převážně HTML/app scraping"},
    {"source": "Burza cenných papírů Praha (PSE)", "type": "exchange news + ZIP", "paywall": "partial", "info_level": 4, "template": "https://www.pse.cz/en/news", "notes": "S2: news HTML + PL.zip (price list), respektovat časová okna"},
    {"source": "ESMA (FIRDS/FITRS)", "type": "EU regulatory datasets", "paywall": "free", "info_level": 5, "template": "https://registers.esma.europa.eu/solr/esma_registers_firds_files/select?q=*&wt=xml&start=0&rows=100", "notes": "S3: SOLR listing + download XML/ZIP, vhodné cache/polling"},
    {"source": "ECB (RSS + MID)", "type": "macro + publications", "paywall": "free", "info_level": 5, "template": "https://www.ecb.europa.eu/rss/press.html", "notes": "S4: oficiální ECB RSS (makro/press/speeches)"},
    {"source": "API info-financiere (FR OAM)", "type": "regulatory API", "paywall": "free", "info_level": 4, "template": "https://info-financiere.gouv.fr/api/explore/v2.1/catalog/datasets", "notes": "S5: otevřené API, limit cca 10k volání/IP/den"},
    {"source": "SEC EDGAR + SEC RSS", "type": "filings + regulator news", "paywall": "free", "info_level": 5, "template": "https://www.sec.gov/news/pressreleases.rss", "notes": "S6: SEC veřejná data; povinný User-Agent u API volání"},
    {"source": "NasdaqTrader Trade Halts", "type": "halts RSS", "paywall": "free", "info_level": 4, "template": "https://www.nasdaqtrader.com/rss.aspx?feed=tradehalts", "notes": "S7: doporučeno nedotazovat častěji než 1x/min"},
    {"source": "NYSE Trade Halts", "type": "halts CSV", "paywall": "free", "info_level": 4, "template": "https://www.nyse.com/api/trade-halts/current/download", "notes": "S8: přímý CSV endpoint (current halts)"},
    {"source": "FINRA Short Volume", "type": "short-sale files", "paywall": "free", "info_level": 4, "template": "https://www.finra.org/finra-data/browse-catalog/short-sale-volume-data", "notes": "S9: veřejné listingy + TXT soubory na CDN"},
    {"source": "GDELT DOC API", "type": "global news aggregator", "paywall": "free", "info_level": 3, "template": "https://api.gdeltproject.org/api/v2/doc/doc?query={ticker}&mode=artlist&format=json", "notes": "S10: API pro monitoring témat/událostí"},
    {"source": "Common Crawl CC-NEWS", "type": "bulk dataset", "paywall": "free", "info_level": 3, "template": "", "notes": "S11: vhodné pro backtesty, vyšší integrační náročnost"},
    {"source": "GlobeNewswire", "type": "press releases", "paywall": "free", "info_level": 3, "template": "https://www.globenewswire.com/RssFeed/orgclass/1/feedTitle/GlobeNewswire%20-%20News%20about%20Public%20Companies", "notes": "S12: RSS/JSON widget feed, může mít throttling"},
    {"source": "PR Newswire", "type": "press releases", "paywall": "free", "info_level": 3, "template": "https://www.prnewswire.com/rss/news-releases-list.rss", "notes": "S13: RSS rozcestník + tematické kanály"},
    {"source": "Business Wire", "type": "press releases", "paywall": "free", "info_level": 3, "template": "https://www.businesswire.com/portal/site/home/news/", "notes": "S14: newsroom/feed options"},
    {"source": "Patria.cz", "type": "financial news CZ", "paywall": "free", "info_level": 3, "template": "https://www.patria.cz/rss.html", "notes": "S15: české ekonomické RSS kanály"},
    {"source": "Akcie.cz", "type": "financial news CZ", "paywall": "partial", "info_level": 2, "template": "https://www.akcie.cz/moje-akcie/rss/", "notes": "S16: RSS kanály, část může vyžadovat login"},
]




def setup_logging() -> logging.Logger:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    return logging.getLogger("refresh_news")


def now_local_naive() -> dt.datetime:
    return dt.datetime.now().replace(tzinfo=None)


def excel_safe(v):
    if isinstance(v, dt.datetime):
        return v.replace(tzinfo=None)
    return v


def slugify_filename(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s)
    return s.strip("_")


def print_bar(prefix: str, i: int, n: int, width: int = 26):
    if n <= 0:
        return
    frac = min(max(i / n, 0.0), 1.0)
    filled = int(round(frac * width))
    bar = "#" * filled + "-" * (width - filled)
    pct = int(round(frac * 100))
    sys.stdout.write(f"\r{prefix} [{bar}] {pct:3d}% ({i}/{n})")
    sys.stdout.flush()
    if i >= n:
        sys.stdout.write("\n")


def load_env_from_code_env(path: str = "code.env") -> Dict[str, str]:
    env = {}
    if not os.path.exists(path):
        return env
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k = k.strip()
            v = v.strip().strip('"').strip("'")
            env[k] = v
            os.environ.setdefault(k, v)
    return env


def mt5_connect():
    try:
        import MetaTrader5 as mt5
    except Exception as e:
        raise RuntimeError("Chybí balík MetaTrader5. Nainstaluj: pip install MetaTrader5") from e

    if not mt5.initialize():
        err = mt5.last_error()
        raise RuntimeError(f"MT5 initialize failed: {err}")

    return mt5


def mt5_visible_symbols(mt5) -> List[str]:
    syms = []
    for s in mt5.symbols_get():
        if getattr(s, "visible", False):
            syms.append(s.name)
    return sorted(set(syms))


def mt5_copy_rates(mt5, symbol: str, timeframe, count: int = 300):
    rates = mt5.copy_rates_from_pos(symbol, timeframe, 0, count)
    return rates


def sma(values: List[float], period: int) -> Optional[float]:
    if len(values) < period:
        return None
    return sum(values[-period:]) / period


def rsi(values: List[float], period: int = 14) -> Optional[float]:
    if len(values) < period + 1:
        return None
    gains = 0.0
    losses = 0.0
    for i in range(-period, 0):
        diff = values[i] - values[i - 1]
        if diff >= 0:
            gains += diff
        else:
            losses -= diff
    if losses == 0:
        return 100.0
    rs = gains / losses
    return 100.0 - (100.0 / (1.0 + rs))


def tech_score_from_mt5(mt5, symbol: str) -> Tuple[Optional[float], Dict[str, Optional[float]], str]:
    tf = mt5.TIMEFRAME_D1
    rates = mt5_copy_rates(mt5, symbol, tf, 250)
    if rates is None or len(rates) < 60:
        return None, {"Close": None, "MA20": None, "MA50": None, "RSI14": None}, "missing"

    closes = [float(r["close"]) for r in rates]
    close = closes[-1]
    ma20 = sma(closes, 20)
    ma50 = sma(closes, 50)
    rsi14 = rsi(closes, 14)

    score = 0.0

    if ma20 is not None and ma50 is not None:
        if ma20 > ma50:
            score += 12
        else:
            score += 4
        if close > ma20:
            score += 8
        else:
            score += 2

    if rsi14 is not None:
        if 45 <= rsi14 <= 65:
            score += 15
        elif 35 <= rsi14 < 45 or 65 < rsi14 <= 75:
            score += 10
        elif 25 <= rsi14 < 35 or 75 < rsi14 <= 85:
            score += 6
        else:
            score += 3

    if ma20 is not None:
        dist = (close - ma20) / ma20 * 100.0
        if dist >= 8:
            score += 2
        elif dist >= 3:
            score += 6
        elif dist >= -3:
            score += 10
        elif dist >= -8:
            score += 6
        else:
            score += 2

    score = max(0.0, min(50.0, score))
    return score, {"Close": close, "MA20": ma20, "MA50": ma50, "RSI14": rsi14}, "ok_mt"




def tech_score_from_yfinance(symbol: str, logger: Optional[logging.Logger] = None) -> Tuple[Optional[float], Dict[str, Optional[float]], str]:
    try:
        import yfinance as yf
    except Exception:
        return None, {"Close": None, "MA20": None, "MA50": None, "RSI14": None}, "missing_yf"

    try:
        hist = yf.Ticker(symbol).history(period="12mo", interval="1d")
        closes = []
        for _, row in hist.iterrows():
            c = row.get("Close")
            if c is None:
                continue
            try:
                closes.append(float(c))
            except Exception:
                continue

        if len(closes) < 60:
            return None, {"Close": None, "MA20": None, "MA50": None, "RSI14": None}, "missing_yf"

        close = closes[-1]
        ma20 = sma(closes, 20)
        ma50 = sma(closes, 50)
        rsi14 = rsi(closes, 14)

        score = 0.0

        if ma20 is not None and ma50 is not None:
            if ma20 > ma50:
                score += 12
            else:
                score += 4
            if close > ma20:
                score += 8
            else:
                score += 2

        if rsi14 is not None:
            if 45 <= rsi14 <= 65:
                score += 15
            elif 35 <= rsi14 < 45 or 65 < rsi14 <= 75:
                score += 10
            elif 25 <= rsi14 < 35 or 75 < rsi14 <= 85:
                score += 6
            else:
                score += 3

        if ma20 is not None:
            dist = (close - ma20) / ma20 * 100.0
            if dist >= 8:
                score += 2
            elif dist >= 3:
                score += 6
            elif dist >= -3:
                score += 10
            elif dist >= -8:
                score += 6
            else:
                score += 2

        score = max(0.0, min(50.0, score))
        return score, {"Close": close, "MA20": ma20, "MA50": ma50, "RSI14": rsi14}, "ok_yf_tech"
    except Exception as e:
        if logger:
            logger.warning("Tech fallback (yfinance) missing for %s: %s", symbol, e)
        return None, {"Close": None, "MA20": None, "MA50": None, "RSI14": None}, "missing_yf"

def yahoo_details_and_score(symbol: str, logger: Optional[logging.Logger] = None) -> Tuple[float, Dict[str, Optional[float]], str]:
    try:
        import yfinance as yf
    except Exception as e:
        raise RuntimeError("Chybí yfinance. Nainstaluj: pip install yfinance") from e

    ysym = symbol
    try:
        t = yf.Ticker(ysym)
        info = getattr(t, "info", {}) or {}

        price = info.get("currentPrice") or info.get("regularMarketPrice") or None
        target = info.get("targetMeanPrice") or info.get("targetMedianPrice") or None
        reco = info.get("recommendationMean") or None
        reco_key = info.get("recommendationKey") or None

        upside = None
        if price and target and price != 0:
            upside = (float(target) / float(price) - 1.0) * 100.0

        score = 0.0
        if reco is not None:
            rm = float(reco)
            if rm <= 1.6:
                score += 8
            elif rm <= 2.2:
                score += 5
            elif rm <= 2.8:
                score += 2
            else:
                score -= 2

        if upside is not None:
            if upside >= 40:
                score += 10
            elif upside >= 20:
                score += 7
            elif upside >= 5:
                score += 3
            elif upside >= -5:
                score += 0
            elif upside >= -15:
                score -= 4
            else:
                score -= 8

        score = max(-20.0, min(20.0, score))

        details = {
            "YahooPrice": float(price) if price is not None else None,
            "YahooTarget": float(target) if target is not None else None,
            "YahooUpsidePct": float(upside) if upside is not None else None,
            "YahooRatingKey": reco_key,
            "YahooRatingMean": float(reco) if reco is not None else None,
        }
        return score, details, "ok_yf"
    except Exception as e:
        if logger:
            logger.warning("Yahoo data missing for %s: %s", symbol, e)
        return 0.0, {"YahooPrice": None, "YahooTarget": None, "YahooUpsidePct": None, "YahooRatingKey": None, "YahooRatingMean": None}, "missing"


@dataclass
class NewsItem:
    ticker: str
    source: str
    title: str
    link: str
    published_utc: Optional[dt.datetime]
    weight: float


def parse_published_dt(entry) -> Optional[dt.datetime]:
    for key in ("published_parsed", "updated_parsed"):
        if hasattr(entry, key) and getattr(entry, key):
            try:
                st = getattr(entry, key)
                return dt.datetime.fromtimestamp(time.mktime(st), tz=dt.timezone.utc)
            except Exception:
                pass
    return None


def source_weight(info_level: int) -> float:
    return 0.5 + (float(info_level) / 5.0) * 1.5




def ticker_candidates(ticker: str) -> List[str]:
    t = (ticker or "").strip().upper()
    if not t:
        return []

    candidates = {t}
    if "." in t:
        candidates.add(t.replace(".", "-"))
        candidates.add(t.replace(".", ""))
    if "-" in t:
        candidates.add(t.replace("-", "."))
        candidates.add(t.replace("-", ""))

    return sorted(candidates)


def entry_mentions_ticker(entry, ticker: str) -> bool:
    text_parts = [
        str(getattr(entry, "title", "") or ""),
        str(getattr(entry, "summary", "") or ""),
        str(getattr(entry, "link", "") or ""),
    ]
    text = " ".join(text_parts).upper()

    for candidate in ticker_candidates(ticker):
        token = candidate.upper()
        if f" {token} " in f" {text} ":
            return True
        if f"({token})" in text or f":{token}" in text or f"/{token}" in text:
            return True
    return False


def fetch_rss_items_for_ticker(
    ticker: str,
    max_per_source: int = 12,
    logger: Optional[logging.Logger] = None,
    source_health: Optional[Dict[str, Dict[str, Any]]] = None,
) -> List[NewsItem]:
    try:
        import feedparser
    except Exception as e:
        raise RuntimeError("Chybí feedparser. Nainstaluj: pip install feedparser") from e

    items: List[NewsItem] = []

    for src in SOURCES:
        source_name = str(src.get("source") or "unknown")
        template = (src.get("template") or "").strip()
        if not template:
            continue

        if source_health is not None:
            state = source_health.setdefault(source_name, {"failures": 0, "disabled": False, "warned_disabled": False})
            if state.get("disabled"):
                if logger and not state.get("warned_disabled"):
                    logger.info("RSS source disabled for this run: %s", source_name)
                    state["warned_disabled"] = True
                continue

        urls = []
        if "{ticker}" in template:
            for c in ticker_candidates(ticker):
                urls.append(template.format(ticker=c))
        else:
            urls.append(template)

        seen_links = set()

        for url in urls:
            try:
                feed = feedparser.parse(url)
            except Exception as e:
                if source_health is not None:
                    state = source_health.setdefault(source_name, {"failures": 0, "disabled": False, "warned_disabled": False})
                    state["failures"] += 1
                    if state["failures"] >= 3:
                        state["disabled"] = True
                if logger:
                    logger.warning("RSS fetch error for %s (%s): %s", ticker, source_name, e)
                continue

            if getattr(feed, "bozo", False) and getattr(feed, "bozo_exception", None):
                if source_health is not None:
                    state = source_health.setdefault(source_name, {"failures": 0, "disabled": False, "warned_disabled": False})
                    state["failures"] += 1
                    if state["failures"] >= 3:
                        state["disabled"] = True
                if logger:
                    logger.warning("RSS parse issue for %s (%s): %s", ticker, source_name, getattr(feed, "bozo_exception", "unknown"))
                continue

            info_level = int(src.get("info_level") or 3)
            w = source_weight(info_level)

            entries = getattr(feed, "entries", []) or []
            taken = 0
            for e in entries:
                if taken >= max_per_source:
                    break

                if "{ticker}" not in template and not entry_mentions_ticker(e, ticker):
                    continue

                title = getattr(e, "title", "") or ""
                link = getattr(e, "link", "") or ""
                if link and link in seen_links:
                    continue
                if link:
                    seen_links.add(link)

                pub = parse_published_dt(e)
                items.append(
                    NewsItem(
                        ticker=ticker,
                        source=str(src.get("source")),
                        title=title,
                        link=link,
                        published_utc=pub,
                        weight=w,
                    )
                )
                taken += 1

    return items


def news_metrics_48h(items: List[NewsItem], now_utc: dt.datetime) -> Tuple[float, int]:
    cutoff = now_utc - dt.timedelta(hours=48)
    wsum = 0.0
    cnt = 0
    for it in items:
        if it.published_utc is None:
            continue
        if it.published_utc >= cutoff:
            cnt += 1
            age_h = (now_utc - it.published_utc).total_seconds() / 3600.0
            recency = max(0.05, 1.0 - (age_h / 48.0))
            wsum += it.weight * recency
    return wsum, cnt


def news_score_0_50(news_weighted_48h: float, news_volume_48h: int) -> float:
    base = 0.0
    base += min(30.0, news_weighted_48h * 4.0)
    base += min(20.0, math.log1p(news_volume_48h) * 5.0)
    return max(0.0, min(50.0, base))


def compute_total_score_macro_logic(news_0_50: float, tech_0_50: float, yahoo_minus20_20: float) -> float:
    total = 0.0
    total += float(news_0_50)
    total += float(tech_0_50)
    total += (float(yahoo_minus20_20) + 20.0) * 0.5
    return max(0.0, min(100.0, total))


def signal_from_total_score_macro_logic(total_0_100: float) -> str:
    if total_0_100 >= 85:
        return "STRONG BUY"
    if total_0_100 >= 70:
        return "BUY"
    if total_0_100 >= 55:
        return "HOLD"
    if total_0_100 >= 40:
        return "SELL"
    return "STRONG SELL"


def _change_pct_between_dates(mt5, symbol: str, start_date: dt.date, end_date: dt.date, yf_period: str = "6mo") -> Tuple[Optional[float], str]:
    def calc_change_from_bars(bars: List[Tuple[dt.date, float, float]]) -> Optional[float]:
        if not bars:
            return None
        bars.sort(key=lambda x: x[0])
        first_day_open = bars[0][1]
        last_day_close = bars[-1][2]
        if first_day_open == 0:
            return None
        return (last_day_close / first_day_open - 1.0) * 100.0

    # 1) Primárně MT5
    mt5_bars: List[Tuple[dt.date, float, float]] = []
    try:
        dt_from = dt.datetime.combine(start_date, dt.time.min)
        dt_to = dt.datetime.combine(end_date + dt.timedelta(days=1), dt.time.min)
        rates = mt5.copy_rates_range(symbol, mt5.TIMEFRAME_D1, dt_from, dt_to)
        if rates is not None:
            for r in rates:
                bar_date = dt.datetime.fromtimestamp(int(r["time"])).date()
                if start_date <= bar_date <= end_date:
                    mt5_bars.append((bar_date, float(r["open"]), float(r["close"])))
    except Exception:
        mt5_bars = []

    change_mt5 = calc_change_from_bars(mt5_bars)
    if change_mt5 is not None:
        return change_mt5, "ok_mt"

    # 2) Fallback: yfinance
    try:
        import yfinance as yf

        hist = yf.Ticker(symbol).history(period=yf_period, interval="1d")
        yf_bars: List[Tuple[dt.date, float, float]] = []
        for idx, row in hist.iterrows():
            bar_date = idx.date()
            if not (start_date <= bar_date <= end_date):
                continue
            o = row.get("Open")
            c = row.get("Close")
            if o is None or c is None:
                continue
            try:
                yf_bars.append((bar_date, float(o), float(c)))
            except Exception:
                continue

        change_yf = calc_change_from_bars(yf_bars)
        if change_yf is not None:
            return change_yf, "ok_yf"
    except Exception:
        pass

    return None, "missing"


def last_week_monday_friday_change_pct(mt5, symbol: str) -> Tuple[Optional[float], str]:
    today = now_local_naive().date()
    this_monday = today - dt.timedelta(days=today.weekday())
    last_monday = this_monday - dt.timedelta(days=7)
    last_friday = last_monday + dt.timedelta(days=4)

    # Některé trhy/symboly nemusí mít svíčku přesně v pondělí/pátek
    # (svátky, specifické obchodní hodiny). Proto se bere první + poslední
    # dostupný obchodní den v okně Po-Pá minulého týdne.
    return _change_pct_between_dates(mt5, symbol, last_monday, last_friday, yf_period="1mo")


def last_1m_change_pct(mt5, symbol: str) -> Tuple[Optional[float], str]:
    end_date = now_local_naive().date()
    start_date = end_date - dt.timedelta(days=30)
    return _change_pct_between_dates(mt5, symbol, start_date, end_date, yf_period="6mo")


def last_3m_change_pct(mt5, symbol: str) -> Tuple[Optional[float], str]:
    end_date = now_local_naive().date()
    start_date = end_date - dt.timedelta(days=90)
    return _change_pct_between_dates(mt5, symbol, start_date, end_date, yf_period="1y")


def try_load_marketcap_map(path: Optional[str]) -> Dict[str, Tuple[Optional[float], Optional[int]]]:
    if path is None:
        for guess in ("market_watch_symbols_enriched_yahoo.xlsx", "market_watch_symbols_enriched_yahoo.csv"):
            if os.path.exists(guess):
                path = guess
                break

    if path is None or not os.path.exists(path):
        return {}

    cap_map: Dict[str, Tuple[Optional[float], Optional[int]]] = {}

    try:
        if path.lower().endswith(".xlsx"):
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            headers = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(1, c).value
                if v:
                    headers[str(v).strip().lower()] = c

            def col(*names):
                for n in names:
                    if n in headers:
                        return headers[n]
                return None

            c_symbol = col("symbol", "yahoo_symbol", "ticker")
            c_cap = col("marketcap_usd", "marketcap", "market_cap")
            c_rank = col("pořadí", "poradi", "rank", "rankmarketcap")

            if not c_symbol:
                return {}

            for r in range(2, ws.max_row + 1):
                sym = ws.cell(r, c_symbol).value
                if not sym:
                    continue
                sym = str(sym).strip()
                cap = ws.cell(r, c_cap).value if c_cap else None
                rk = ws.cell(r, c_rank).value if c_rank else None
                try:
                    cap_f = float(cap) if cap is not None and cap != "" else None
                except Exception:
                    cap_f = None
                try:
                    rk_i = int(rk) if rk is not None and rk != "" else None
                except Exception:
                    rk_i = None
                cap_map[sym] = (cap_f, rk_i)

        else:
            import csv
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    sym = (row.get("symbol") or row.get("yahoo_symbol") or row.get("ticker") or "").strip()
                    if not sym:
                        continue
                    cap = row.get("marketcap_usd") or row.get("marketcap") or ""
                    rk = row.get("poradi") or row.get("pořadí") or row.get("rank") or row.get("rankmarketcap") or ""
                    cap_f = None
                    rk_i = None
                    try:
                        cap_f = float(cap) if cap else None
                    except Exception:
                        cap_f = None
                    try:
                        rk_i = int(float(rk)) if rk else None
                    except Exception:
                        rk_i = None
                    cap_map[sym] = (cap_f, rk_i)

    except Exception:
        return {}

    return cap_map


def create_workbook_template() -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "Signals"

    headers = [
        "Ticker",
        "MT5Symbol",
        "UpdatedAt",
        "MarketCapUSD",
        "RankMarketCap",
        "NewsWeighted48h",
        "NewsVolume48h",
        "NewsScore(0-50)",
        "TechScore(0-50)",
        "YahooScore(-20..20)",
        "TotalScore(0-100)",
        "Signal",
        "TechStatus",
        "YahooStatus",
        "Close",
        "MA20",
        "MA50",
        "RSI14",
        "YahooPrice",
        "YahooTarget",
        "YahooUpsidePct",
        "YahooRatingKey",
        "YahooRatingMean",
        "LastWeekMonFriChangePct",
        "LastWeekMonFriDropPctText",
        "LastWeekMonFriStatus",
        "Last1MChangePct",
        "Last1MChangePctText",
        "Last1MStatus",
        "Last3MChangePct",
        "Last3MChangePctText",
        "Last3MStatus",
    ]

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True)
        ws.cell(1, c).alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    ws_src = wb.create_sheet("Sources")
    ws_src.append(["Source", "Type", "Paywall", "InfoLevel(1-5)", "RSS/Feed template (optional)", "Notes"])
    for c in range(1, 7):
        ws_src.cell(1, c).font = Font(bold=True)
    for s in SOURCES:
        ws_src.append([s["source"], s["type"], s["paywall"], s["info_level"], s["template"], s["notes"]])
    ws_src.freeze_panes = "A2"

    ws_art = wb.create_sheet("Articles")
    ws_art.append(["Ticker", "Source", "PublishedUTC", "Title", "Link", "Weight"])
    for c in range(1, 7):
        ws_art.cell(1, c).font = Font(bold=True)
    ws_art.freeze_panes = "A2"

    return wb


# =========================
# ✅ ADDED: Dashboard builder
# =========================
def build_dashboard(wb: Workbook):
    # vytvoř / vyčisti sheet
    if "Dashboard" in wb.sheetnames:
        ws = wb["Dashboard"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Dashboard")

    ws["A1"] = "GeneratedAt"
    ws["B1"] = now_local_naive().strftime("%Y-%m-%d %H:%M:%S")
    ws["A1"].font = Font(bold=True)

    # načti data ze Signals
    ws_sig = wb["Signals"]
    headers = [ws_sig.cell(1, c).value for c in range(1, ws_sig.max_column + 1)]
    idx = {str(h): i for i, h in enumerate(headers) if h}

    def get(row, name, default=None):
        col = idx.get(name)
        if col is None:
            return default
        return row[col]

    data = []
    for r in range(2, ws_sig.max_row + 1):
        row = [ws_sig.cell(r, c).value for c in range(1, ws_sig.max_column + 1)]
        ticker = get(row, "Ticker")
        mcap = get(row, "MarketCapUSD")
        if not ticker:
            continue
        data.append({
            "Ticker": ticker,
            "MarketCapUSD": mcap,
            "RankMarketCap": get(row, "RankMarketCap"),
            "TotalScore": get(row, "TotalScore(0-100)"),
            "Signal": get(row, "Signal"),
            "NewsScore": get(row, "NewsScore(0-50)"),
            "TechScore": get(row, "TechScore(0-50)"),
            "YahooScore": get(row, "YahooScore(-20..20)"),
            "LastWeekMonFriChangePct": get(row, "LastWeekMonFriChangePct"),
            "LastWeekMonFriStatus": get(row, "LastWeekMonFriStatus"),
            "Last1MChangePct": get(row, "Last1MChangePct"),
            "Last1MStatus": get(row, "Last1MStatus"),
            "Last3MChangePct": get(row, "Last3MChangePct"),
            "Last3MStatus": get(row, "Last3MStatus"),
        })

    def write_section(title, start_row, cols, rows):
        ws[f"A{start_row}"] = title
        ws[f"A{start_row}"].font = Font(bold=True)
        header_row = start_row + 1

        for i, h in enumerate(cols, 1):
            cell = ws.cell(header_row, i, h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for j, rr in enumerate(rows, 1):
            for i, h in enumerate(cols, 1):
                ws.cell(header_row + j, i, rr.get(h))

        return header_row + len(rows) + 2

    # Top 20 by TotalScore
    top_total = sorted(
        [d for d in data if d["TotalScore"] is not None],
        key=lambda x: float(x["TotalScore"]),
        reverse=True
    )[:20]
    top_total_rows = []
    for i, d in enumerate(top_total, 1):
        top_total_rows.append({
            "Rank": i,
            "Ticker": d["Ticker"],
            "TotalScore(0-100)": d["TotalScore"],
            "Signal": d["Signal"],
            "MarketCapUSD": d["MarketCapUSD"],
            "NewsScore(0-50)": d["NewsScore"],
            "TechScore(0-50)": d["TechScore"],
            "YahooScore(-20..20)": d["YahooScore"],
        })

    r = 3
    r = write_section(
        "Top 20 by TotalScore",
        r,
        ["Rank", "Ticker", "TotalScore(0-100)", "Signal", "MarketCapUSD", "NewsScore(0-50)", "TechScore(0-50)", "YahooScore(-20..20)"],
        top_total_rows
    )

    # Největší 20 propadů za minulý týden (pondělí -> pátek)
    biggest_weekly_drops = sorted(
        [d for d in data if d["LastWeekMonFriChangePct"] is not None and float(d["LastWeekMonFriChangePct"]) < 0],
        key=lambda x: float(x["LastWeekMonFriChangePct"])
    )[:20]
    biggest_weekly_drops_rows = []
    for i, d in enumerate(biggest_weekly_drops, 1):
        biggest_weekly_drops_rows.append({
            "Rank": i,
            "Ticker": d["Ticker"],
            "LastWeekMonFriChangePct": f"{float(d['LastWeekMonFriChangePct']):.2f}%",
            "TotalScore(0-100)": d["TotalScore"],
            "Signal": d["Signal"],
            "LastWeekMonFriStatus": d["LastWeekMonFriStatus"],
        })

    r = write_section(
        "Top 20 nejvetsi propady minuly tyden (Po-Pa, %)",
        r,
        ["Rank", "Ticker", "LastWeekMonFriChangePct", "TotalScore(0-100)", "Signal", "LastWeekMonFriStatus"],
        biggest_weekly_drops_rows
    )

    # Největší 20 propadů za poslední měsíc
    biggest_1m_drops = sorted(
        [d for d in data if d["Last1MChangePct"] is not None and float(d["Last1MChangePct"]) < 0],
        key=lambda x: float(x["Last1MChangePct"])
    )[:20]
    biggest_1m_drops_rows = []
    for i, d in enumerate(biggest_1m_drops, 1):
        biggest_1m_drops_rows.append({
            "Rank": i,
            "Ticker": d["Ticker"],
            "Last1MChangePct": f"{float(d['Last1MChangePct']):.2f}%",
            "TotalScore(0-100)": d["TotalScore"],
            "Signal": d["Signal"],
            "Last1MStatus": d["Last1MStatus"],
        })

    r = write_section(
        "Top 20 nejvetsi propady za 1 mesic (%)",
        r,
        ["Rank", "Ticker", "Last1MChangePct", "TotalScore(0-100)", "Signal", "Last1MStatus"],
        biggest_1m_drops_rows
    )

    # Největší 20 propadů za poslední 3 měsíce
    biggest_3m_drops = sorted(
        [d for d in data if d["Last3MChangePct"] is not None and float(d["Last3MChangePct"]) < 0],
        key=lambda x: float(x["Last3MChangePct"])
    )[:20]
    biggest_3m_drops_rows = []
    for i, d in enumerate(biggest_3m_drops, 1):
        biggest_3m_drops_rows.append({
            "Rank": i,
            "Ticker": d["Ticker"],
            "Last3MChangePct": f"{float(d['Last3MChangePct']):.2f}%",
            "TotalScore(0-100)": d["TotalScore"],
            "Signal": d["Signal"],
            "Last3MStatus": d["Last3MStatus"],
        })

    r = write_section(
        "Top 20 nejvetsi propady za 3 mesice (%)",
        r,
        ["Rank", "Ticker", "Last3MChangePct", "TotalScore(0-100)", "Signal", "Last3MStatus"],
        biggest_3m_drops_rows
    )

    # Top 20 by MarketCap
    top_mcap = sorted(
        [d for d in data if d["MarketCapUSD"] is not None],
        key=lambda x: float(x["MarketCapUSD"]),
        reverse=True
    )[:20]
    top_mcap_rows = []
    for i, d in enumerate(top_mcap, 1):
        top_mcap_rows.append({
            "Rank": i,
            "Ticker": d["Ticker"],
            "MarketCapUSD": d["MarketCapUSD"],
            "RankMarketCap": d["RankMarketCap"],
            "TotalScore(0-100)": d["TotalScore"],
            "Signal": d["Signal"],
        })

    r = write_section(
        "Top 20 by MarketCap",
        r,
        ["Rank", "Ticker", "MarketCapUSD", "RankMarketCap", "TotalScore(0-100)", "Signal"],
        top_mcap_rows
    )

    # ✅ Bottom 20 by MarketCap
    bottom_mcap = sorted(
        [d for d in data if d["MarketCapUSD"] is not None],
        key=lambda x: float(x["MarketCapUSD"])
    )[:20]
    bottom_mcap_rows = []
    for i, d in enumerate(bottom_mcap, 1):
        bottom_mcap_rows.append({
            "Rank": i,
            "Ticker": d["Ticker"],
            "MarketCapUSD": d["MarketCapUSD"],
            "RankMarketCap": d["RankMarketCap"],
            "TotalScore(0-100)": d["TotalScore"],
            "Signal": d["Signal"],
        })

    write_section(
        "Bottom 20 by MarketCap",
        r,
        ["Rank", "Ticker", "MarketCapUSD", "RankMarketCap", "TotalScore(0-100)", "Signal"],
        bottom_mcap_rows
    )


def main():
    logger = setup_logging()

    parser = argparse.ArgumentParser()
    parser.add_argument("--outdir", default=".", help="Kam uložit výstupní Excel (default: aktuální složka)")
    parser.add_argument("--marketcap", default=None, help="Volitelný soubor s marketcap/rank (xlsx/csv)")
    args = parser.parse_args()

    env = load_env_from_code_env("code.env")
    if env:
        logger.info("ENV: loaded %s vars from code.env", len(env))

    outdir = args.outdir
    os.makedirs(outdir, exist_ok=True)

    cap_map = try_load_marketcap_map(args.marketcap)
    if cap_map:
        logger.info("MarketCap: loaded %s symbols from file", len(cap_map))

    print("Step 1/4: MT5 watchlist symbols ...")
    mt5 = mt5_connect()
    symbols = mt5_visible_symbols(mt5)
    print(f"MT5 symbols (visible): {len(symbols)}")

    print("Step 2/4: Collect RSS news (supported sources only) ...")
    now_utc = dt.datetime.now(dt.timezone.utc)
    all_items: Dict[str, List[NewsItem]] = {}
    source_health: Dict[str, Dict[str, Any]] = {}
    n = len(symbols)
    for i, sym in enumerate(symbols, 1):
        print_bar("RSS", i, n)
        try:
            items = fetch_rss_items_for_ticker(sym, max_per_source=10, logger=logger, source_health=source_health)
        except RuntimeError as e:
            logger.error("RSS unavailable for %s: %s", sym, e)
            items = []
        all_items[sym] = items

    print("Step 3/4: Compute signals (news + tech + yahoo + marketcap/rank) ...")
    wb = create_workbook_template()
    ws = wb["Signals"]
    ws_art = wb["Articles"]

    for i, sym in enumerate(symbols, 1):
        print_bar("Signals", i, n)

        items = all_items.get(sym, []) or []
        for it in items[:50]:
            pub = it.published_utc.isoformat() if it.published_utc else ""
            ws_art.append([it.ticker, it.source, pub, it.title, it.link, it.weight])

        news_w48, news_v48 = news_metrics_48h(items, now_utc)
        news_score = news_score_0_50(news_w48, news_v48)

        tech_score, tech_details, tech_status = tech_score_from_mt5(mt5, sym)
        if tech_score is None:
            tech_score, tech_details, tech_status = tech_score_from_yfinance(sym, logger=logger)
        if tech_score is None:
            tech_score = 0.0

        yahoo_score, ydetails, ystatus = yahoo_details_and_score(sym, logger=logger)
        last_week_drop_pct, last_week_drop_status = last_week_monday_friday_change_pct(mt5, sym)
        last_1m_drop_pct, last_1m_drop_status = last_1m_change_pct(mt5, sym)
        last_3m_drop_pct, last_3m_drop_status = last_3m_change_pct(mt5, sym)

        total = compute_total_score_macro_logic(news_score, tech_score, yahoo_score)
        signal = signal_from_total_score_macro_logic(total)

        cap, rank = cap_map.get(sym, (None, None))

        updated = now_local_naive().strftime("%Y-%m-%d %H:%M")

        ws.append([
            sym,
            sym,
            updated,
            cap,
            rank,
            round(news_w48, 3),
            int(news_v48),
            round(news_score, 1),
            round(tech_score, 1),
            round(yahoo_score, 1),
            round(total, 1),
            signal,
            tech_status,
            ystatus,
            tech_details.get("Close"),
            tech_details.get("MA20"),
            tech_details.get("MA50"),
            tech_details.get("RSI14"),
            ydetails.get("YahooPrice"),
            ydetails.get("YahooTarget"),
            ydetails.get("YahooUpsidePct"),
            ydetails.get("YahooRatingKey"),
            ydetails.get("YahooRatingMean"),
            round(last_week_drop_pct, 2) if last_week_drop_pct is not None else None,
            f"{last_week_drop_pct:.2f}%" if last_week_drop_pct is not None else None,
            last_week_drop_status,
            round(last_1m_drop_pct, 2) if last_1m_drop_pct is not None else None,
            f"{last_1m_drop_pct:.2f}%" if last_1m_drop_pct is not None else None,
            last_1m_drop_status,
            round(last_3m_drop_pct, 2) if last_3m_drop_pct is not None else None,
            f"{last_3m_drop_pct:.2f}%" if last_3m_drop_pct is not None else None,
            last_3m_drop_status,
        ])

    print("\nStep 4/4: SAVE workbook ...")

    # =========================
    # ✅ ADDED: Build Dashboard
    # =========================
    build_dashboard(wb)

    ts = now_local_naive().strftime("%Y%m%d_%H%M%S")
    outname = f"market_checker_watchlist_{ts}.xlsx"
    outpath = os.path.join(outdir, outname)
    outpath = os.path.abspath(outpath)
    wb.save(outpath)

    print(f"OK: saved -> {outpath}")


if __name__ == "__main__":
    raise SystemExit(main())
